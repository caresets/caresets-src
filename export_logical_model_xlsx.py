"""
Build a logical-model workbook in the "Logical Model Template" layout from the
FHIR StructureDefinitions in input/models/ and the glossary mappings in
input/glossary_mappings.csv.

Each file carries the template's Instructions tab plus one data sheet named
after the model, styled from the template's "Data Set 1" tab:

  A  Transaction / process     the model (name - title)
  B  Name                      element path below the root, e.g. administeredProduct.lotNumber
  C  Short Label FR            element.short, French
  D  Description FR            element.definition, French
  E  Short Label NL            element.short, Dutch
  F  Description NL            element.definition, Dutch
  G  Short Label EN            element.short, English
  H  Description EN            element.definition, English
  I  Data Type                 element.type[*].code, joined with " | "
  J  min occurrence            element.min  (0, 1, ...)      see --cardinality
  K  max occurrence            element.max  (1, *, ...)      see --cardinality
  L  ValueSet                  the last segment of element.binding.valueSet
  M  Code        (glossary)    GlossaryCode from the mappings CSV
  N  Relationship(glossary)    --relationship, only where a code was matched
  O  Example Value             left empty - not held in the models
  P  Example Value Display     left empty

A model's text is carried in FHIR as a base value plus `translation` extensions
on the matching `_short` / `_definition` element. This reads whichever language
is in the base field and whichever are in the extensions, and puts each in its
own column, so an author never has to see the encoding.

Each file also gets a "Model" sheet holding the model's own title and
description per language, which is what the viewer shows at the top of a page.

One .xlsx per model, written into models/xls/ by default.

Usage:
  python export_logical_model_xlsx.py
  python export_logical_model_xlsx.py --model BeModelVaccination BeModelProblem
  python export_logical_model_xlsx.py --include-draft --valueset-sheets
"""

import argparse
import csv
import glob
import json
import os
import sys
import warnings

warnings.filterwarnings("ignore")
import openpyxl  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

ROOT = os.path.dirname(os.path.abspath(__file__))
TEMPLATE = os.path.join(ROOT, "Logical Model Template.xlsx")
PROTOTYPE = "Data Set 1"          # the tab whose header and styling we clone
KEEP = ("Instructions",)          # template tabs carried into the output
FIRST_DATA_ROW = 3
COLS = {"process": 1, "name": 2,
        "short_fr": 3, "description_fr": 4,
        "short_nl": 5, "description_nl": 6,
        "short_en": 7, "description_en": 8,
        "datatype": 9, "min": 10, "max": 11, "valueset": 12,
        "binding_strength": 13, "code": 14, "relationship": 15}
HEADERS = [
    ("Transaction / process", ""),
    ("Data Element", "Name"),
    ("", "Short Label FR"), ("", "Description FR"),
    ("", "Short Label NL"), ("", "Description NL"),
    ("", "Short Label EN"), ("", "Description EN"),
    ("", "Data Type"), ("", "min occurrence"), ("", "max occurrence"), ("", "ValueSet"),
    ("", "Binding Strength"),
    ("Common Glossary", "Code"), ("", "Relationship"),
    ("Example Value", ""), ("Example Value Display Name (for coded elements)", ""),
]
LANGS = ("fr", "nl", "en")
TRANSLATION_URL = "http://hl7.org/fhir/StructureDefinition/translation"
# Excel caps a cell at 32767 characters; long FHIR definitions can approach it.
CELL_LIMIT = 32000


def load_mappings(path):
    """(model filename, element suffix) -> glossary code."""
    out = {}
    if not os.path.exists(path):
        return out
    with open(path, encoding="utf-8-sig", newline="") as fh:
        for r in csv.DictReader(fh, delimiter=";"):
            model = (r.get("Model") or "").strip()
            suffix = (r.get("ElementSuffix") or "").strip()
            code = (r.get("GlossaryCode") or "").strip()
            if model and suffix and code:
                out[(model, suffix)] = code
    return out


def model_keys(doc, filename):
    """Every identifier a mapping row may legitimately use for this model.

    A model's identity is its `name` and `url`, which live inside the file;
    the filename does not survive a rename or a re-export, so it is accepted
    only for compatibility with older mapping files."""
    return {str(k).strip().lower()
            for k in (doc.get("name"), doc.get("url"), doc.get("id"), filename) if k}


def match_code(mappings, keys, suffix):
    """add_glossary_mappings.py matches on endswith('.' + suffix), so mirror
    that here rather than requiring the suffix to be spelled out in full."""
    for (model, mapped), code in mappings.items():
        if model.strip().lower() in keys and (
            suffix == mapped or suffix.endswith("." + mapped)
        ):
            return code
    return None


def cell(value):
    if value is None:
        return None
    text = str(value)
    return text[:CELL_LIMIT] + " [truncated]" if len(text) > CELL_LIMIT else text


def occurrence(element, style):
    """FHIR min and max as they stand - 0 and *, so the pair reads 0..* across
    the two columns. The template's own example instead uses y/n flags
    (mandatory? repeating?), which --cardinality yn reproduces."""
    lo, hi = element.get("min"), element.get("max")
    if style == "numeric":
        return (lo if lo is not None else ""), (hi if hi is not None else "")
    is_mandatory = isinstance(lo, int) and lo >= 1
    repeats = hi == "*" or (str(hi).isdigit() and int(hi) > 1)
    return ("y" if is_mandatory else "n"), ("y" if repeats else "n")


def elements_of(doc):
    """Every differential element below the root, in document order, deduped."""
    root = doc.get("type") or doc.get("name") or ""
    rows, seen = [], set()
    for e in doc.get("differential", {}).get("element", []):
        path = e.get("path", "")
        if "." not in path:
            continue                      # the root element itself
        suffix = path.split(".", 1)[1]
        if suffix in seen:
            continue
        seen.add(suffix)
        rows.append((suffix, e))
    if not rows:                          # fall back for snapshot-only models
        for e in doc.get("snapshot", {}).get("element", []):
            path = e.get("path", "")
            if "." not in path:
                continue
            suffix = path.split(".", 1)[1]
            if suffix in seen:
                continue
            seen.add(suffix)
            rows.append((suffix, e))
    del root
    return rows


def sheet_title(name, used):
    """Excel: 31 chars, no []:*?/\\ , and unique in the workbook."""
    clean = "".join(ch for ch in name if ch not in "[]:*?/\\")[:31] or "Model"
    title, n = clean, 2
    while title in used:
        suffix = "_%d" % n
        title = clean[: 31 - len(suffix)] + suffix
        n += 1
    used.add(title)
    return title


def texts(owner, field, base_lang):
    """{lang: text} for one field, from the base value plus its translations.

    FHIR keeps a primitive's extensions on a sibling starting with "_", so
    `definition` is translated by `_definition`; each translation extension
    pairs a `lang` code with its `content`.
    """
    out = {}
    base = owner.get(field)
    if base:
        out[base_lang] = base
    sibling = owner.get("_" + field) or {}
    for ext in sibling.get("extension", []) or []:
        if ext.get("url") != TRANSLATION_URL:
            continue
        lang = content = None
        for sub in ext.get("extension", []) or []:
            if sub.get("url") == "lang":
                lang = sub.get("valueCode") or sub.get("valueString")
            elif sub.get("url") == "content":
                content = sub.get("valueString")
        if lang and content:
            out[lang] = content
    return out


def base_language(doc):
    """The language the untranslated fields are written in.

    `language` on the resource says so explicitly. The eHealth exports do not
    set it and are written in English, which is the fallback.
    """
    return (doc.get("language") or "en").split("-")[0].lower()


def valueset_name(element):
    vs = (element.get("binding") or {}).get("valueSet")
    if not vs:
        return None
    return vs.rstrip("/").split("/")[-1].split("|")[0]


def write_model_sheet(wb, proto, doc, filename, mappings, args, used_titles):
    ws = wb.copy_worksheet(proto)
    ws.title = sheet_title(doc.get("name") or filename, used_titles)

    # The prototype carries the template's own example rows; clear everything
    # below the two header rows but keep their styling.
    if ws.max_row >= FIRST_DATA_ROW:
        ws.delete_rows(FIRST_DATA_ROW, ws.max_row - FIRST_DATA_ROW + 1)

    label = doc.get("name") or filename
    if doc.get("title") and doc["title"] != label:
        label = "%s - %s" % (label, doc["title"])
    if doc.get("version"):
        label = "%s (v%s)" % (label, doc["version"])

    # The template's own two header rows describe a single-language sheet;
    # rewrite them for the per-language columns, keeping their styling.
    from copy import copy as _copy
    style_top = _copy(ws.cell(row=1, column=2)._style)
    style_sub = _copy(ws.cell(row=2, column=2)._style)
    for i, (top, sub) in enumerate(HEADERS, start=1):
        c1, c2 = ws.cell(row=1, column=i), ws.cell(row=2, column=i)
        c1.value, c2.value = (top or None), (sub or None)
        c1._style, c2._style = _copy(style_top), _copy(style_sub)
    for i in range(1, len(HEADERS) + 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width =             46 if i in (3, 4, 5, 6, 7, 8) else 22

    base = base_language(doc)
    mapped = 0
    keys = model_keys(doc, filename)
    for i, (suffix, e) in enumerate(elements_of(doc)):
        r = FIRST_DATA_ROW + i
        lo, hi = occurrence(e, args.cardinality)
        code = match_code(mappings, keys, suffix)
        if code:
            mapped += 1
        short = texts(e, "short", base)
        definition = texts(e, "definition", base)
        values = {
            "process": label,
            "name": suffix,
            "datatype": " | ".join(t.get("code", "") for t in e.get("type", []) if t.get("code")) or None,
            "min": lo,
            "max": hi,
            "valueset": valueset_name(e),
            "binding_strength": (e.get("binding") or {}).get("strength"),
            "code": code,
            "relationship": args.relationship if code else None,
        }
        for lang in LANGS:
            values["short_" + lang] = cell(short.get(lang))
            values["description_" + lang] = cell(definition.get(lang))
        for key, col in COLS.items():
            ws.cell(row=r, column=col).value = values[key]
    return ws, mapped


def write_model_sheet_meta(wb, doc, used_titles):
    """A small sheet for the model's own title and description per language -
    what the viewer shows at the top of the page, and what an author needs to
    be able to edit without touching JSON."""
    ws = wb.create_sheet(sheet_title("Model", used_titles), 0)
    base = base_language(doc)
    title = texts(doc, "title", base)
    description = texts(doc, "description", base)
    rows = [
        ("Model name", doc.get("name")),
        ("Canonical URL", doc.get("url")),
        ("Version", doc.get("version")),
        ("Status", doc.get("status")),
        ("Base language", base),
        (None, None),
        ("Title FR", title.get("fr")), ("Description FR", description.get("fr")),
        ("Title NL", title.get("nl")), ("Description NL", description.get("nl")),
        ("Title EN", title.get("en")), ("Description EN", description.get("en")),
    ]
    for i, (k, v) in enumerate(rows, start=1):
        ws.cell(row=i, column=1).value = k
        ws.cell(row=i, column=2).value = v
        if k:
            ws.cell(row=i, column=1).font = openpyxl.styles.Font(bold=True)
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 90
    return ws


def write_valueset_sheets(wb, proto_vs, docs, used_titles):
    """One tab per distinct binding, in the template's ValueSet layout. The
    codes themselves are left for a terminologist - the models bind to a
    ValueSet URL but do not carry its expansion."""
    seen = {}
    for doc, filename in docs:
        for suffix, e in elements_of(doc):
            vs = (e.get("binding") or {}).get("valueSet")
            if vs:
                seen.setdefault(valueset_name(e), (vs, doc.get("name"), suffix))
    for name, (url, model, suffix) in sorted(seen.items()):
        ws = wb.copy_worksheet(proto_vs)
        ws.title = sheet_title(name, used_titles)
        ws.cell(row=1, column=2).value = name
        ws.cell(row=2, column=2).value = "%s.%s" % (model, suffix)
        ws.cell(row=3, column=2).value = url
        if ws.max_row >= 6:
            ws.delete_rows(6, ws.max_row - 5)
    return len(seen)


def safe_filename(name):
    """A model name is already tame, but Windows still forbids these."""
    return "".join(ch for ch in name if ch not in '<>:"/|?*' + chr(92)).strip() or "model"


def build_workbook(doc, filename, mappings, args):
    """One workbook holding one model, built fresh from the template so each
    file keeps the template's Instructions tab and styling."""
    wb = openpyxl.load_workbook(args.template)
    if PROTOTYPE not in wb.sheetnames:
        sys.exit("Template has no %r tab (has: %s)" % (PROTOTYPE, wb.sheetnames))
    proto = wb[PROTOTYPE]
    proto_vs = next((wb[n] for n in wb.sheetnames if n.startswith("VS")), None)

    used = set(wb.sheetnames)
    ws, mapped = write_model_sheet(wb, proto, doc, filename, mappings, args, used)
    write_model_sheet_meta(wb, doc, used)
    rows = ws.max_row - FIRST_DATA_ROW + 1

    vs_count = 0
    if args.valueset_sheets and proto_vs is not None:
        vs_count = write_valueset_sheets(wb, proto_vs, [(doc, filename)], used)

    keep = set(KEEP) | {ws.title} | {n for n in wb.sheetnames if n.startswith("Model")}
    for name in list(wb.sheetnames):
        if name in keep:
            continue
        if name == PROTOTYPE or name.startswith("Data Set") or (
            proto_vs is not None and name == proto_vs.title
        ):
            del wb[name]
    return wb, rows, mapped, vs_count


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--models", default=os.path.join("input", "models"),
                    help="folder of StructureDefinition JSON (default: input/models)")
    ap.add_argument("--model", nargs="+", metavar="NAME",
                    help="only these models, by their StructureDefinition `name` "
                         "(default: every model found)")
    ap.add_argument("--include-draft", action="store_true",
                    help="also include models under the draft/ subfolder")
    ap.add_argument("--mappings", default=os.path.join("input", "glossary_mappings.csv"))
    ap.add_argument("--template", default=TEMPLATE)
    ap.add_argument("--out-dir", dest="out_dir", default=os.path.join("models", "xls"),
                    help="folder to write one <ModelName>.xlsx per model into "
                         "(default: models/xls)")
    ap.add_argument("--cardinality", choices=("numeric", "yn"), default="numeric",
                    help="'numeric' writes the FHIR min and max as they stand, so the two "
                         "columns read 0..* ; 'yn' instead writes the y/n flags the template's "
                         "own example row uses - mandatory? repeating? (default: numeric)")
    ap.add_argument("--relationship", default="equivalent",
                    help="value for the Common Glossary Relationship column on mapped rows "
                         "(default: equivalent); pass '' to leave it blank")
    ap.add_argument("--valueset-sheets", action="store_true",
                    help="add one tab per ValueSet the model binds to, in the template's "
                         "ValueSet layout, for a terminologist to fill in")
    args = ap.parse_args()

    if not os.path.exists(args.template):
        sys.exit("Template not found: %s" % args.template)

    pattern = "**/*.json" if args.include_draft else "*.json"
    paths = sorted(glob.glob(os.path.join(args.models, pattern), recursive=args.include_draft))
    if not paths:
        sys.exit("No models found under %s" % args.models)

    docs = []
    for p in paths:
        with open(p, encoding="utf-8") as fh:
            doc = json.load(fh)
        if doc.get("resourceType") != "StructureDefinition":
            continue
        if args.model and doc.get("name") not in args.model:
            continue
        docs.append((doc, os.path.basename(p)))
    if not docs:
        sys.exit("No StructureDefinition matched. Asked for: %s" % (args.model or "all"))

    mappings = load_mappings(args.mappings)
    os.makedirs(args.out_dir, exist_ok=True)

    made = []
    for doc, filename in sorted(docs, key=lambda d: d[0].get("name") or ""):
        wb, rows, mapped, vs_count = build_workbook(doc, filename, mappings, args)
        out = os.path.join(args.out_dir, safe_filename(doc.get("name") or filename) + ".xlsx")
        wb.save(out)
        made.append((os.path.basename(out), rows, mapped, vs_count))

    sys.stdout.reconfigure(encoding="utf-8")
    print("Template: %s" % args.template)
    print("Written : %d file(s) into %s" % (len(made), args.out_dir))
    print("cardinality style '%s'" % args.cardinality)
    total_rows = total_mapped = 0
    for name, rows, mapped, vs_count in made:
        total_rows += rows
        total_mapped += mapped
        print("  %-40s %3d element(s), %3d mapped%s"
              % (name, rows, mapped, ", %d ValueSet tab(s)" % vs_count if vs_count else ""))
    print("  %-40s %3d element(s), %3d mapped (%.0f%%)"
          % ("TOTAL", total_rows, total_mapped,
             100.0 * total_mapped / total_rows if total_rows else 0))
    unmapped = [n for n, _, m, _ in made if m == 0]
    if unmapped:
        print("\n  ! no glossary mapping at all in %d file(s): %s"
              % (len(unmapped), ", ".join(unmapped)))


if __name__ == "__main__":
    main()
