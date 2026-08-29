#!/usr/bin/env python
"""Extract Active glossary definitions (FR/NL from the workbook, EN from the CSVs)
into a single JSON payload for definition-quality review.

Usage:
    python extract_definitions.py [--workbook PATH] [--sheet NAME]
                                  [--status Active] [--out PATH]

Writes JSON to --out (default: stdout).
"""
import argparse
import csv
import glob
import json
import os
import sys
import warnings

warnings.filterwarnings("ignore")
import openpyxl  # noqa: E402

HERE = os.path.dirname(os.path.abspath(__file__))
REPO = os.path.abspath(os.path.join(HERE, "..", "..", "..", ".."))

# Column headers we look for in the workbook sheet, mapped to output keys.
# Matching is case-insensitive and accent-insensitive on the header text.
COLUMN_ALIASES = {
    "caresets": "careset",
    "statut de la def": "status",
    "item": "term",
    "synonym": "synonym",
    "synonyme": "synonym",
    "suggested min": "min",
    "suggested max": "max",
    "definition fr": "def_fr",
    "description fr": "desc_fr",
    "definition nl": "def_nl",
    "description nl": "desc_nl",
    "definition en": "def_en",
    "description en": "desc_en",
    "datatype": "datatype",
}


def norm(s):
    """Lowercase, strip, and fold accents so header matching is robust."""
    if s is None:
        return ""
    import unicodedata

    s = unicodedata.normalize("NFKD", str(s))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return " ".join(s.lower().split())


def clean(v):
    if v is None:
        return None
    s = str(v).replace("\xa0", " ").strip()
    return s or None


def find_workbook(explicit):
    if explicit:
        return explicit
    pattern = os.path.join(REPO, "input", "Glossaire CareSets*.xlsx")
    hits = sorted(glob.glob(pattern), key=os.path.getmtime, reverse=True)
    if not hits:
        sys.exit("No workbook found matching %s" % pattern)
    return hits[0]


def load_en_from_csvs():
    """Term -> {source, en, fr, nl, status} from the published glossary CSVs."""
    out = {}
    for name in ("ClinicalGlossary.csv", "OperationalGlossary.csv"):
        path = os.path.join(REPO, "input", name)
        if not os.path.exists(path):
            continue
        with open(path, encoding="utf-8-sig", newline="") as fh:
            for row in csv.DictReader(fh, delimiter=";"):
                term = clean(row.get("Term"))
                if not term:
                    continue
                out.setdefault(
                    norm(term),
                    {
                        "csv_source": name,
                        "csv_status": clean(row.get("Status")),
                        "csv_en": clean(row.get("EN")),
                        "csv_fr": clean(row.get("FR")),
                        "csv_nl": clean(row.get("NL")),
                    },
                )
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--workbook")
    ap.add_argument("--sheet", default="Glossaire v1")
    ap.add_argument("--status", default="Active")
    ap.add_argument("--out")
    args = ap.parse_args()

    wb_path = find_workbook(args.workbook)
    wb = openpyxl.load_workbook(wb_path, data_only=True)
    if args.sheet not in wb.sheetnames:
        sys.exit("Sheet %r not in %s (have: %s)" % (args.sheet, wb_path, wb.sheetnames))
    ws = wb[args.sheet]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        sys.exit("Sheet %r is empty" % args.sheet)
    header = rows[0]
    idx = {}
    for i, h in enumerate(header):
        key = COLUMN_ALIASES.get(norm(h))
        if key and key not in idx:
            idx[key] = i

    missing = [k for k in ("status", "term", "def_fr") if k not in idx]
    if missing:
        sys.exit("Sheet %r is missing required column(s): %s" % (args.sheet, missing))

    en_map = load_en_from_csvs()
    want = norm(args.status)
    entries = []
    for rownum, r in enumerate(rows[1:], start=2):
        if norm(r[idx["status"]]) != want:
            continue
        e = {"row": rownum}
        for key, i in idx.items():
            e[key] = clean(r[i]) if i < len(r) else None
        e.update(en_map.get(norm(e.get("term")), {}))
        entries.append(e)

    payload = {
        "workbook": os.path.relpath(wb_path, REPO).replace("\\", "/"),
        "sheet": args.sheet,
        "status_filter": args.status,
        "columns_found": sorted(idx),
        "count": len(entries),
        "entries": entries,
    }
    text = json.dumps(payload, ensure_ascii=False, indent=2)
    if args.out:
        with open(args.out, "w", encoding="utf-8") as fh:
            fh.write(text)
        print("Wrote %d entries to %s" % (len(entries), args.out))
    else:
        sys.stdout.reconfigure(encoding="utf-8")
        print(text)


if __name__ == "__main__":
    main()
