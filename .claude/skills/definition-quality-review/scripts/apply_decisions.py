#!/usr/bin/env python
"""Apply accepted / revised definition-quality findings to a copy of the workbook.

Inputs:
  --findings   the sidecar JSON written alongside the review report
               (list of {id, row, term, proposed: {fr, nl, en}})
  --decisions  the JSON produced by collect_decisions.py --json
  --workbook   source .xlsx (default: newest input/Glossaire CareSets*.xlsx)
  --out        destination .xlsx (REQUIRED; never overwrites the source)

Only findings whose decision is 'accept' (use the proposed text) or 'revise'
(use the reviewer's own text) are written. 'reject' and 'pending' are skipped.
EN has no column in the workbook, so EN changes are reported as a CSV patch
for input/ClinicalGlossary.csv rather than written to the .xlsx.
"""
import argparse
import glob
import json
import os
import sys
import warnings

warnings.filterwarnings("ignore")
import openpyxl  # noqa: E402
from openpyxl.styles import Color, PatternFill  # noqa: E402

HERE = os.path.dirname(os.path.abspath(__file__))
REPO = os.path.abspath(os.path.join(HERE, "..", "..", "..", ".."))
LANG_COL = {"fr": "Definition FR", "nl": "Définition NL"}
ITEM_COL = "Item"
STATUS_COL = "statut de la def"


def norm(s):
    import unicodedata

    s = unicodedata.normalize("NFKD", str(s or ""))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return " ".join(s.lower().split())


def fill_rgb(cell):
    """The cell's solid fill colour as 'FFRRGGBB', or None. Theme and indexed
    colours come back as non-strings, which is why this is not a one-liner."""
    f = cell.fill
    if f is None or f.patternType is None:
        return None
    rgb = getattr(f.fgColor, "rgb", None)
    return rgb if isinstance(rgb, str) else None


def mark_font(cell, rgb):
    """Recolour a cell's font, keeping bold/italic/size/name as they were."""
    from copy import copy

    font = copy(cell.font)
    font.color = Color(rgb=rgb)
    cell.font = font


def restyle_status_fill(ws, col, status, rgb):
    """Keep the status highlight in the status column alone: the fill on rows
    carrying `status`, and that same fill cleared everywhere else it appears.

    Row-level fills matter as much as cell ones - Excel paints them edge to
    edge across the whole row, and openpyxl does not move them when rows are
    inserted, so a stale one ends up marking the wrong term.
    """
    target = rgb if len(rgb) == 8 else "FF" + rgb
    green = PatternFill(fill_type="solid", fgColor=target)
    clear = PatternFill(fill_type=None)
    set_n = cleared = rows_cleared = 0
    for r in range(2, ws.max_row + 1):
        is_status = norm(ws.cell(row=r, column=col["status"]).value) == norm(status)
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            if c == col["status"] and is_status:
                cell.fill = green
                set_n += 1
            elif fill_rgb(cell) == target:
                cell.fill = clear
                cleared += 1
    for r, dim in list(ws.row_dimensions.items()):
        f = getattr(dim, "fill", None)
        if f is not None and f.patternType:
            got = getattr(f.fgColor, "rgb", None)
            if isinstance(got, str) and got == target:
                dim.fill = clear
                rows_cleared += 1
    return set_n, cleared, rows_cleared


def insert_new_term(ws, col, term, texts, status):
    """Add a row for a term that has none yet, alphabetically inside the block
    of rows already carrying `status`. Returns the row number used.

    Call this only after every cell edit is done: inserting shifts the rows
    below it, which would invalidate the row numbers held in the sidecar.
    """
    block = [r for r in range(2, ws.max_row + 1)
             if norm(ws.cell(row=r, column=col["status"]).value) == norm(status)]
    at = ws.max_row + 1
    if block:
        at = block[-1] + 1
        for r in block:
            if norm(ws.cell(row=r, column=col["item"]).value) > norm(term):
                at = r
                break
    ws.insert_rows(at)
    ws.cell(row=at, column=col["item"]).value = term
    ws.cell(row=at, column=col["status"]).value = status
    for lang, text in texts.items():
        if lang in col and text:
            ws.cell(row=at, column=col[lang]).value = text
    return at


def write_report(path, source, out, applied, inserted, en_patch, skipped):
    o = []
    o.append("# Applied glossary changes\n")
    o.append("- Source workbook: `%s`" % source)
    o.append("- Written to: `%s`" % out)
    o.append("- Cells changed: **%d** / Rows added: **%d** / EN rows to patch: **%d**\n"
             % (len(applied), len(inserted), len(en_patch)))
    o.append("The source workbook is not modified. EN has no column in the sheet, so English "
             "changes are listed below for `input/ClinicalGlossary.csv` and are not written "
             "here.\n")

    if inserted:
        o.append("## Rows added\n")
        for i in inserted:
            o.append("### `%s` - new row %s\n" % (i["term"], i["row"]))
            for lang in ("fr", "nl"):
                if i["texts"].get(lang):
                    o.append("- **%s:** %s" % (lang.upper(), i["texts"][lang]))
            o.append("")

    if applied:
        o.append("## Cells changed\n")
        last = object()
        for a in applied:
            if a["term"] != last:
                last = a["term"]
                o.append("### `%s` (row %s)\n" % (a["term"], a.get("row_out", a["row"])))
            o.append("**%s** - %s\n" % (a["lang"], a["finding"]))
            o.append("| | |")
            o.append("|--|--|")
            o.append("| Before | %s |" % ((a["before"] or "*(empty)*").replace("|", "\\|").replace("\n", " ")))
            o.append("| After | %s |" % (a["after"].replace("|", "\\|").replace("\n", " ")))
            o.append("")

    if en_patch:
        o.append("## EN changes for `input/ClinicalGlossary.csv`\n")
        o.append("| Term | Finding | New EN definition |")
        o.append("|------|---------|-------------------|")
        for p in en_patch:
            o.append("| %s | %s | %s |" % (p["term"], p["finding"], p["en"].replace("|", "\\|")))
        o.append("")

    if skipped:
        o.append("## Not applied by this run (%d)\n" % len(skipped))
        o.append("| Finding | Reason |")
        o.append("|---------|--------|")
        for fid, reason in skipped:
            o.append("| %s | %s |" % (fid, reason.replace("|", "\\|")))
        o.append("")

    with open(path, "w", encoding="utf-8", newline="\n") as fh:
        fh.write("\n".join(o))


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--findings", required=True)
    ap.add_argument("--decisions", required=True)
    ap.add_argument("--workbook")
    ap.add_argument("--sheet", default="Glossaire v1")
    ap.add_argument("--out", required=True)
    ap.add_argument("--add-new-terms", dest="add_new", metavar="STATUS", nargs="?", const="Active",
                    help="also add a row for each decided term that has no row in the sheet yet, "
                         "with this status (default Active). Without it such findings are skipped.")
    ap.add_argument("--report", help="write a human-readable change report (.md) as well")
    ap.add_argument("--mark-changes", dest="mark", nargs="?", const="800000", metavar="RRGGBB",
                    help="colour the font of every cell this run writes or adds, so new text is "
                         "visible against what was already in the sheet (default maroon 800000)")
    ap.add_argument("--status-fill-only", dest="status_fill", nargs="?", const="92D050",
                    metavar="RRGGBB",
                    help="keep the status highlight in the status column alone - set it on rows "
                         "carrying the --add-new-terms status (default Active) and clear that same "
                         "fill from every other cell (default green 92D050)")
    args = ap.parse_args()

    wb_path = args.workbook
    if not wb_path:
        hits = sorted(
            glob.glob(os.path.join(REPO, "input", "Glossaire CareSets*.xlsx")),
            key=os.path.getmtime,
            reverse=True,
        )
        if not hits:
            sys.exit("No workbook found in input/")
        wb_path = hits[0]
    if os.path.abspath(wb_path) == os.path.abspath(args.out):
        sys.exit("--out must differ from the source workbook")

    findings = {f["id"]: f for f in json.load(open(args.findings, encoding="utf-8"))}
    decisions = json.load(open(args.decisions, encoding="utf-8"))["decisions"]

    wb = openpyxl.load_workbook(wb_path)
    ws = wb[args.sheet]
    header = [c.value for c in ws[1]]
    col = {}
    for key, name in list(LANG_COL.items()) + [("item", ITEM_COL), ("status", STATUS_COL)]:
        for i, h in enumerate(header, start=1):
            if norm(h) == norm(name):
                col[key] = i
                break

    applied, skipped, en_patch, new_terms = [], [], [], {}
    for d in decisions:
        f = findings.get(d["id"])
        if f is None:
            skipped.append((d["id"], "no such finding id in sidecar"))
            continue
        if d["status"] not in ("accept", "revise"):
            skipped.append((d["id"], d["status"]))
            continue
        if f.get("lang") == "meta" or not f.get("row"):
            # No cell to write: either a cross-cutting finding, or a new term
            # that has no row in the sheet yet. Text left on one is inert, so
            # say so loudly rather than dropping it on the floor.
            if f.get("lang") == "meta":
                note = "cross-cutting finding - apply by hand"
                if any(d.get(k) for k in ("fr", "nl", "en")):
                    note += ("; text supplied here will NOT be written, put it on the "
                             "finding for that row/language")
            elif args.add_new:
                # A decided term with no row yet: collect its languages now and
                # insert one row per term once all cell edits are done.
                bucket = new_terms.setdefault(f["term"], {"texts": {}, "findings": []})
                bucket["findings"].append(d["id"])
                proposed = f.get("proposed") or {}
                for lang in ("fr", "nl", "en"):
                    text = d.get(lang) if d["status"] == "revise" else None
                    if text is None and d["status"] == "accept":
                        text = proposed.get(lang)
                    if not text:
                        continue
                    if lang == "en":
                        en_patch.append({"term": f["term"], "en": text, "finding": d["id"]})
                    else:
                        bucket["texts"][lang] = text
                continue
            else:
                note = ("new term '%s' - it has no row in the sheet yet; re-run with "
                        "--add-new-terms to add it" % f.get("term"))
            skipped.append((d["id"], note))
            continue
        proposed = f.get("proposed") or {}
        for lang in ("fr", "nl", "en"):
            text = d.get(lang) if d["status"] == "revise" else None
            if text is None and d["status"] == "accept":
                text = proposed.get(lang)
            if not text:
                continue
            if lang == "en":
                en_patch.append({"term": f["term"], "en": text, "finding": d["id"]})
                continue
            if lang not in col:
                skipped.append((d["id"], "no %s column in sheet" % lang.upper()))
                continue
            cell = ws.cell(row=f["row"], column=col[lang])
            applied.append(
                {
                    "finding": d["id"],
                    "row": f["row"],
                    "term": f["term"],
                    "lang": lang.upper(),
                    "before": cell.value,
                    "after": text,
                }
            )
            cell.value = text

    # Rows are inserted only after every edit above, because inserting shifts
    # the rows below it and the sidecar's row numbers would no longer hold.
    inserted = []
    for term in sorted(new_terms):
        bucket = new_terms[term]
        at = insert_new_term(ws, col, term, bucket["texts"], args.add_new)
        inserted.append({"term": term, "row": at, "texts": bucket["texts"],
                         "findings": bucket["findings"], "status": args.add_new})

    # Every insert shifts the rows at or below it down by one, so a source row
    # number no longer addresses the same cell in the output. Resolve each edit
    # to its row in the written workbook before touching it again or reporting it.
    for a in applied:
        a["row_out"] = a["row"] + sum(1 for i in inserted if i["row"] <= a["row"])

    restyled = None
    if args.status_fill:
        restyled = restyle_status_fill(ws, col, args.add_new or "Active", args.status_fill)

    if args.mark:
        rgb = args.mark if len(args.mark) == 8 else "FF" + args.mark
        for a in applied:
            mark_font(ws.cell(row=a["row_out"], column=col[a["lang"].lower()]), rgb)
        for i in inserted:
            for lang in i["texts"]:
                mark_font(ws.cell(row=i["row"], column=col[lang]), rgb)
            mark_font(ws.cell(row=i["row"], column=col["item"]), rgb)

    wb.save(args.out)

    sys.stdout.reconfigure(encoding="utf-8")
    print("Source : %s" % wb_path)
    print("Written: %s" % args.out)
    print("Applied %d cell change(s), added %d row(s); skipped %d finding(s)."
          % (len(applied), len(inserted), len(skipped)))
    for a in applied:
        print("  row %-4s %-22s %s" % (a.get("row_out", a["row"]), a["term"], a["lang"]))
    for i in inserted:
        print("  + row %-2s %-22s new term (%s)" % (i["row"], i["term"], i["status"]))
    if en_patch:
        print("\nEN changes (no EN column in the workbook) - patch input/ClinicalGlossary.csv:")
        for p in en_patch:
            print("  %-22s %s" % (p["term"], p["en"]))
    log = os.path.splitext(args.out)[0] + ".changelog.json"
    with open(log, "w", encoding="utf-8") as fh:
        json.dump(
            {"source": wb_path, "out": args.out, "applied": applied, "inserted": inserted,
             "en_patch": en_patch, "skipped": skipped},
            fh, ensure_ascii=False, indent=2,
        )
    if args.mark:
        print("Changed and added text set in #%s." % args.mark)
    if restyled:
        print("Status highlight: set on %d status cell(s), cleared from %d other cell(s) "
              "and %d whole-row fill(s)." % restyled)
    print("\nChangelog: %s" % log)
    if args.report:
        write_report(args.report, wb_path, args.out, applied, inserted, en_patch, skipped)
        print("Report   : %s" % args.report)


if __name__ == "__main__":
    main()
