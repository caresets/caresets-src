"""
Rebuild `Logical Model Template.xlsx`, the workbook given to model editors.

The template and the generator have to agree column for column: an editor who
fills in a column `import_logical_model_xlsx.py` does not read loses the text
silently. Generating the template from the same column definition the generator
uses is what keeps them from drifting.

Shape, following the original template:

  Instructions      how to fill it in
  Model             one row per data set - the model's own name, title and
                    description in each language
  Data Set 1..n     one sheet per model, one row per data element
  VS<Name>          one sheet per ValueSet the models bind to

The existing allergy example is preserved as worked content: its English text
moves into the `Short Label EN` / `Description EN` columns, and an editor sees
a filled-in row rather than an empty grid.

Usage:
  python make_logical_model_template.py
  python make_logical_model_template.py --out "Logical Model Template.xlsx"
"""

import argparse
import os
import shutil
import sys
import warnings

warnings.filterwarnings("ignore")
import openpyxl  # noqa: E402
from openpyxl.styles import Alignment, Font  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

ROOT = os.path.dirname(os.path.abspath(__file__))
DEFAULT_OUT = "Logical Model Template.xlsx"

# (row 1 group heading, row 2 column heading, width). The row-2 headings are
# what import_logical_model_xlsx.py matches on, so they are the contract.
COLUMNS = [
    ("Transaction / process", "",                 26),
    ("Data Element",          "Name",             26),
    ("",                      "Short Label FR",   34),
    ("",                      "Description FR",   46),
    ("",                      "Short Label NL",   34),
    ("",                      "Description NL",   46),
    ("",                      "Short Label EN",   34),
    ("",                      "Description EN",   46),
    ("",                      "Data Type",        16),
    ("",                      "min occurrence",   15),
    ("",                      "max occurrence",   15),
    ("",                      "ValueSet",         24),
    ("Common Glossary",       "Code",             22),
    ("",                      "Relationship",     16),
    ("Example Value",         "",                 20),
    ("Example Value Display Name (for coded values)", "", 26),
]

MODEL_COLUMNS = [
    ("Data Set sheet", 18), ("Model name", 26), ("Canonical URL", 52),
    ("Version", 10), ("Status", 12),
    ("Title FR", 34), ("Description FR", 52),
    ("Title NL", 34), ("Description NL", 52),
    ("Title EN", 34), ("Description EN", 52),
]

INSTRUCTIONS = [
    ("role", "Tab", "action"),
    ("Data modeler / BA / SME", "Model",
     "One row per data set. Name the model, and give its title and description "
     "in each language you have. Leave a language blank if you do not have it."),
    ("Data modeler / BA / SME", "Data Set X",
     "Rename the tab to the model it describes, then list one data element per row."),
    ("Data modeler / BA / SME", "Data Set X, column A",
     "The use case or interaction the element belongs to."),
    ("Data modeler / BA / SME", "Data Set X, column B",
     "The element name, as it will appear in the model. Use a dot for nesting, "
     "e.g. administeredProduct.lotNumber."),
    ("Data modeler / BA / SME", "Data Set X, columns C-H",
     "Short label and description per language. Write the languages you have; "
     "English is added and verified before publication. Where a language is "
     "missing the published model falls back to French, then Dutch."),
    ("Data modeler / BA / SME", "Data Set X, columns I-K",
     "Data type, and whether the element is mandatory and repeating. "
     "min/max are the FHIR cardinality: 0 or 1, and 1 or *."),
    ("Data modeler / BA / SME", "Data Set X, column L",
     "For coded elements, the ValueSet name. Add a sheet named VS<Name> for it."),
    ("Terminologist", "Data Set X, columns M-N",
     "Map the element to a Common Glossary concept: the concept code, and how "
     "it relates to this element. Leave blank if there is no matching concept."),
    ("Terminologist", "VS<Name>",
     "The codes allowed for a coded element, with the code system they come from."),
]


def style_header(cell, bold=True, wrap=False):
    cell.font = Font(bold=bold)
    cell.alignment = Alignment(vertical="top", wrap_text=wrap)


def build_instructions(wb):
    ws = wb.create_sheet("Instructions", 0)
    for r, row in enumerate(INSTRUCTIONS, start=1):
        for c, value in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if r == 1:
                style_header(cell)
    for col, width in (("A", 24), ("B", 26), ("C", 96)):
        ws.column_dimensions[col].width = width
    return ws


def build_model_sheet(wb, models):
    ws = wb.create_sheet("Model", 1)
    for c, (heading, width) in enumerate(MODEL_COLUMNS, start=1):
        style_header(ws.cell(row=1, column=c, value=heading))
        ws.column_dimensions[get_column_letter(c)].width = width
    for r, model in enumerate(models, start=2):
        for c, (heading, _) in enumerate(MODEL_COLUMNS, start=1):
            ws.cell(row=r, column=c).value = model.get(heading)
    ws.freeze_panes = "A2"
    return ws


def build_data_set(wb, title, rows):
    ws = wb.create_sheet(title)
    for c, (group, heading, width) in enumerate(COLUMNS, start=1):
        style_header(ws.cell(row=1, column=c, value=group or None))
        style_header(ws.cell(row=2, column=c, value=heading or None), bold=False, wrap=True)
        ws.column_dimensions[get_column_letter(c)].width = width
    # Group headings span their columns, as in the original template.
    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=12)
    ws.merge_cells(start_row=1, start_column=13, end_row=1, end_column=14)
    for r, row in enumerate(rows, start=3):
        for c, (_, heading, _) in enumerate(COLUMNS, start=1):
            key = heading or ("process" if c == 1 else
                              "example" if c == 15 else "example_display")
            ws.cell(row=r, column=c).value = row.get(key)
    ws.freeze_panes = "A3"
    return ws


def build_valueset(wb, name, context, codes):
    ws = wb.create_sheet(name)
    style_header(ws.cell(row=1, column=1, value="ValueSet"))
    ws.cell(row=1, column=2, value=name)
    style_header(ws.cell(row=2, column=1, value="Context"))
    ws.cell(row=2, column=2, value=context)
    style_header(ws.cell(row=3, column=1, value="FHIR context:"))
    headings = ["System (if known)", "System Version", "Code", "Short Label",
                "Description", "Short Label FR", "Description FR",
                "Short Label NL", "Description NL"]
    for c, h in enumerate(headings, start=1):
        style_header(ws.cell(row=5, column=c, value=h), wrap=True)
        ws.column_dimensions[get_column_letter(c)].width = 20 if c < 3 else 30
    for r, code in enumerate(codes, start=6):
        for c, value in enumerate(code, start=1):
            ws.cell(row=r, column=c).value = value
    return ws


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--out", default=DEFAULT_OUT)
    args = ap.parse_args()

    out = args.out if os.path.isabs(args.out) else os.path.join(ROOT, args.out)
    if os.path.exists(out):
        shutil.copy2(out, out + ".bak")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    build_instructions(wb)

    # The allergy example from the original template, kept as worked content so
    # an editor sees a filled row rather than an empty grid. Its English text
    # moves into the EN columns.
    build_model_sheet(wb, [
        {"Data Set sheet": "Data Set 1", "Model name": "BeModelAllergy",
         "Canonical URL": "https://www.ehealth.fgov.be/standards/fhir/core-clinical/"
                          "StructureDefinition/BeModelAllergy",
         "Version": "0.1.0", "Status": "draft",
         "Title EN": "Allergy Model",
         "Description EN": "Logical model of the CareSet Allergy."},
        {"Data Set sheet": "Data Set 2", "Model name": "", "Canonical URL": "",
         "Version": "", "Status": "draft"},
    ])

    build_data_set(wb, "Data Set 1", [
        {"process": "T1 - Submit Allergy", "Name": "RecordedDate",
         "Short Label EN": "Date of registering the allergy",
         "Description EN": "Date of registering the allergy. This is the date the "
                           "record was created, not the date of onset.",
         "Data Type": "date", "min occurrence": "1", "max occurrence": "1",
         "Code": "RecordedDate", "Relationship": "equivalent",
         "example": "2022-01-01"},
        {"process": "T1 - Submit Allergy", "Name": "AllergyType",
         "Short Label EN": "The type of risk - allergy, intolerance or non-allergic "
                           "hypersensitivity",
         "Description EN": "The type of risk the record describes.",
         "Data Type": "Coded", "min occurrence": "0", "max occurrence": "1",
         "ValueSet": "VSAllergyType", "example": "1", "example_display": "Allergy"},
    ])
    build_data_set(wb, "Data Set 2", [])

    build_valueset(wb, "VSAllergyType", "AllergyType", [
        ("SNOMED", None, "1", "Allergy", None, "Allergie", None, "Allergie", None),
        ("SNOMED", None, "2", "Intolerance", None, "Intolérance", None, "Intolerantie", None),
        ("SNOMED", None, "3", "Non-allergic hypersensitivity", None,
         "Hypersensibilité non allergique", None, "Niet-allergische overgevoeligheid", None),
    ])

    wb.save(out)
    sys.stdout.reconfigure(encoding="utf-8")
    print("Wrote %s" % args.out)
    print("  sheets: %s" % ", ".join(wb.sheetnames))
    print("  element columns: %s" % ", ".join(h for _, h, _ in COLUMNS if h))
    return 0


if __name__ == "__main__":
    sys.exit(main())
