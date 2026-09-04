"""
Merge confirmed glossary mappings from the CSV back into the model workbooks.

Step 4 of the mapping flow:

  1  published StructureDefinitions  ->  models/xls/*.xlsx
  2  propose_model_mappings.py writes proposals into input/glossary_mappings.csv
  3  a reviewer sets Status to confirmed or rejected
  4  THIS SCRIPT writes the confirmed ones into the workbooks' Code column

Only `confirmed` rows are merged. A row with no Status counts as confirmed -
the mappings written before the Status column existed are all decided ones.
`proposed` and `rejected` rows are left alone, so re-running after a partial
review picks up exactly what has been decided since.

The workbook is where a mapping is authored, so this closes the loop: the CSV
carries a proposal until somebody confirms it, and confirmation puts it in the
source.

  python merge_mappings_to_xlsx.py --dry-run    show what would change
  python merge_mappings_to_xlsx.py              write the workbooks
  python merge_mappings_to_xlsx.py --force      also overwrite differing codes
"""

import argparse
import csv
import glob
import io
import os
import sys
import warnings

warnings.filterwarnings("ignore")
import openpyxl  # noqa: E402

import glossary_terms  # noqa: E402

ROOT = os.path.dirname(os.path.abspath(__file__))
BOOKS = os.path.join("models", "xls")
MAPPINGS = os.path.join("input", "glossary_mappings.csv")
RELATIONSHIP = "equivalent"


def confirmed_rows(path, allow_unapproved=False):
    """(model, element suffix) -> glossary code, for confirmed rows only.

    A confirmed row whose target is not an approved glossary term is held back
    and reported. The GlossaryStatus column in the CSV is not trusted for this:
    it is recomputed here from the glossary itself, so a term withdrawn since
    the row was written is caught rather than waved through.
    """
    terms = glossary_terms.load()
    drafted = glossary_terms.workbook_terms()
    out, skipped, unapproved = {}, {}, []
    with io.open(path, encoding="utf-8-sig", newline="") as fh:
        for r in csv.DictReader(fh, delimiter=";"):
            model = (r.get("Model") or "").strip()
            suffix = (r.get("ElementSuffix") or "").strip()
            code = (r.get("GlossaryCode") or "").strip()
            if not (model and suffix and code):
                continue
            # No Status is the pre-Status-column shape, and those are decided.
            status = (r.get("Status") or "confirmed").strip().lower()
            if status != "confirmed":
                skipped[status] = skipped.get(status, 0) + 1
                continue
            target = glossary_terms.status_of(code, terms, drafted)
            if target != "approved":
                unapproved.append((model, suffix, code, target))
                if not allow_unapproved:
                    continue
            out[(model, suffix)] = code
    return out, skipped, unapproved


def workbook_paths():
    """model name -> workbook path. The Model sheet is authoritative for the
    name; the filename only agrees with it by convention."""
    books = {}
    for path in sorted(glob.glob(os.path.join(ROOT, BOOKS, "*.xlsx"))):
        name = os.path.splitext(os.path.basename(path))[0]
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        if "Model" in wb.sheetnames:
            m = wb["Model"]
            for row in m.iter_rows(min_row=1, max_row=m.max_row, max_col=2):
                if str(row[0].value or "").strip().lower() == "model name" and row[1].value:
                    name = str(row[1].value).strip()
                    break
        wb.close()
        books[name] = path
    return books


def merge(mappings, books, force=False, dry_run=False):
    written, unchanged, conflicts, missing = 0, 0, [], []
    touched = set()

    by_model = {}
    for (model, suffix), code in mappings.items():
        by_model.setdefault(model, {})[suffix] = code

    for model, wanted in sorted(by_model.items()):
        path = books.get(model)
        if path is None:
            missing.append((model, sorted(wanted)))
            continue
        wb = openpyxl.load_workbook(path)
        changed = 0
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            if ws.max_row < 3:
                continue
            h = {str(ws.cell(2, c).value or "").strip(): c
                 for c in range(1, ws.max_column + 1)}
            if "Name" not in h or "Code" not in h:
                continue
            for r in range(3, ws.max_row + 1):
                element = ws.cell(r, h["Name"]).value
                if not element:
                    continue
                # Match the way add_glossary_mappings.py does, so a mapping
                # means the same thing here as it does downstream: the key is
                # either the element's full path or a trailing part of it.
                suffix = str(element).strip()
                code = wanted.get(suffix)
                if not code:
                    code = next((c for k, c in wanted.items()
                                 if suffix.endswith("." + k)), None)
                if not code:
                    continue
                current = ws.cell(r, h["Code"]).value
                current = str(current).strip() if current else ""
                if current == code:
                    unchanged += 1
                    continue
                if current and not force:
                    conflicts.append((model, str(element).strip(), current, code))
                    continue
                if not dry_run:
                    ws.cell(r, h["Code"]).value = code
                    if h.get("Relationship") and not ws.cell(r, h["Relationship"]).value:
                        ws.cell(r, h["Relationship"]).value = RELATIONSHIP
                changed += 1
        if changed:
            if not dry_run:
                wb.save(path)
            written += changed
            touched.add(model)
            print("  %-34s %2d mapping(s)" % (model, changed))
        wb.close()
    return written, unchanged, conflicts, missing, touched


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--mappings", default=MAPPINGS)
    ap.add_argument("--dry-run", action="store_true",
                    help="report what would change, without writing")
    ap.add_argument("--allow-unapproved", action="store_true",
                    help="also write mappings whose target is not an approved "
                         "glossary term; without this they are reported and held "
                         "back, because a mapping to a term the glossary does not "
                         "publish will not resolve on the site")
    ap.add_argument("--force", action="store_true",
                    help="overwrite a Code that differs from the confirmed one; "
                         "without this a difference is reported, not resolved")
    args = ap.parse_args()

    sys.stdout.reconfigure(encoding="utf-8")
    path = args.mappings if os.path.isabs(args.mappings) else os.path.join(ROOT, args.mappings)
    if not os.path.exists(path):
        sys.exit("No mappings CSV at %s" % args.mappings)

    mappings, skipped, unapproved = confirmed_rows(path, args.allow_unapproved)
    print("Confirmed  : %d mapping(s)" % len(mappings))
    for status, n in sorted(skipped.items()):
        print("             %-10s %3d row(s) left alone" % (status, n))
    if unapproved:
        print("\n%d confirmed row(s) do not point at an approved glossary term%s:"
              % (len(unapproved), "" if args.allow_unapproved else " - held back"))
        for model, suffix, code, target in unapproved[:15]:
            print("  %-28s %-24s %-22s %s" % (model, suffix, code, target))
        if len(unapproved) > 15:
            print("  ... and %d more" % (len(unapproved) - 15))
        print("Either approve the term in the glossary, point the mapping at a "
              "term that is\napproved, or pass --allow-unapproved to write them "
              "anyway.")
    if not mappings:
        print("\nNothing confirmed yet. Set Status to confirmed in %s first." % args.mappings)
        return 0

    books = workbook_paths()
    print("Workbooks  : %d in %s\n" % (len(books), BOOKS))

    written, unchanged, conflicts, missing, touched = merge(
        mappings, books, force=args.force, dry_run=args.dry_run)

    print("\n%s %d mapping(s) across %d workbook(s); %d already present"
          % ("Would write" if args.dry_run else "Wrote", written, len(touched), unchanged))

    if conflicts:
        print("\n%d element(s) already carry a different code - left as they are."
              % len(conflicts))
        print("The workbook is the source, so a difference is a real disagreement, "
              "not a stale value. Re-run with --force only if the CSV is right.")
        for model, element, current, code in conflicts[:15]:
            print("  %-28s %-24s %s -> %s" % (model, element, current, code))
        if len(conflicts) > 15:
            print("  ... and %d more" % (len(conflicts) - 15))

    if missing:
        print("\n%d model(s) in the CSV have no workbook:" % len(missing))
        for model, suffixes in missing[:10]:
            print("  %-34s %d mapping(s)" % (model, len(suffixes)))

    if not args.dry_run and written:
        print("\nNext: python import_logical_model_xlsx.py   "
              "(regenerates the models from the workbooks)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
