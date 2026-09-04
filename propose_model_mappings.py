"""
Propose glossary mappings for model elements, into the mappings CSV.

Only 38 of 602 model elements carry a glossary code. This proposes the rest,
including the ones a name match would miss - `subject` is the Patient,
`basedOn` is the OriginatingRequest, `bodySite` is the BodyLocation.

It is step 2 of four:

  1  fetch_ehealth_models.py  ->  import_models_zip.py  ->  export_logical_model_xlsx.py
     published StructureDefinitions become models/xls/*.xlsx
  2  THIS SCRIPT proposes mappings into input/glossary_mappings.csv as
     Status=proposed, alongside the confirmed ones already there
  3  a reviewer opens that CSV and sets Status to confirmed or rejected
  4  merge_mappings_to_xlsx.py writes the confirmed ones back into the
     workbooks, which stay the place a mapping is authored

Nothing downstream acts on a proposed row: add_glossary_mappings.py and
export_logical_model_xlsx.py both skip anything not confirmed, so a proposal
sitting in the CSV cannot reach the site or the StructureDefinitions by
accident. A blank Status means confirmed, which is what the rows written before
this column existed are.

Each proposed row carries its confidence, the reason, and the model's own
description of the element, so the CSV can be reviewed in Excel without having
to open the model beside it.

  python propose_model_mappings.py             propose
  python propose_model_mappings.py --report    also write a grouped read-through
"""


import argparse
import glob
import io
import csv
import csv
import json
import os
import re
import shutil
import sys
import time
import warnings

warnings.filterwarnings("ignore")
import openpyxl  # noqa: E402

import glossary_terms  # noqa: E402

ROOT = os.path.dirname(os.path.abspath(__file__))
BOOKS = os.path.join("models", "xls")
MAPPINGS = os.path.join("input", "glossary_mappings.csv")
MAPPINGS = os.path.join("input", "glossary_mappings.csv")
OUT_DIR = "glossary-changes"
FIELDS = ["Model", "ElementSuffix", "GlossaryCode", "GlossaryStatus", "Status",
          "Confidence", "Rationale", "ElementDescription"]
FIELDS = ["Model", "ElementSuffix", "GlossaryCode", "GlossaryStatus", "Status",
          "Confidence", "Rationale", "ElementDescription"]
RELATIONSHIP = "equivalent"

# element name (lower-cased) -> (glossary code, confidence, why, nested?)
#
# `nested` says whether the rule may apply below the top level. It is False for
# the CareSet's own fields: Code and Status name the concept and lifecycle of
# the record itself, so section.code and adherence.status are not those things.
# That was decided on 29 August - "leave all six unmapped" - and a rule that
# ignores it would put the decision back on the reviewer every time.
#
# "certain"  the element is that concept under another name; the description
#            confirms it in every model where it appears
# "likely"   the mapping holds in the models seen, but the name is generic
#            enough that a model could use it differently
# "check"    plausible and worth a decision, but read the descriptions first
RULES = {
    # --- identity, provenance, lifecycle ------------------------------------
    "identifier": ("BusinessIdentifier", "certain", "the CareSet instance's business identifier", False),
    "businessidentifier": ("BusinessIdentifier", "certain", "named for the concept", False),
    "patient": ("Patient", "certain", "the person the record is about", False),
    "subject": ("Patient", "likely", "the person the record is about, under FHIR's generic name; check any "
                "model where the subject could be an organisation or a device", False),
    "recorder": ("Recorder", "certain", "the professional who records the content", False),
    "author": ("Recorder", "certain", "Author is the synonym of Recorder since decision F-012; the descriptions "
               "say 'the person who encodes'", False),
    "asserter": ("Asserter", "certain", "the person who reports the information", False),
    "performer": ("Performer", "certain", "the person who performed the act", False),
    "recordeddate": ("RecordedDate", "certain", "when the record was entered", False),
    "recorded": ("RecordedDate", "certain", "when the record was entered", False),
    "creationdate": ("RecordedDate", "likely", "'creation date ... will not change' is the date the record was "
                     "entered, which is what RecordedDate names", False),
    "status": ("Status", "likely", "the record's lifecycle state; exclude any model where status means "
               "clinical currency, which is ClinicalStatus", False),
    "clinicalstatus": ("ClinicalStatus", "certain", "named for the concept", False),
    "verificationstatus": ("VerificationStatus", "certain", "named for the concept", False),

    # --- content ------------------------------------------------------------
    "code": ("Code", "certain", "the coded clinical concept the record refers to", False),
    "category": ("Category", "certain", "classifies the element by clinical meaning", False),
    "note": ("Note", "certain", "free-text additional information", True),

    # --- relations between records -----------------------------------------
    "partof": ("PartOf", "certain", "the encompassing CareSet", False),
    "basedon": ("OriginatingRequest", "likely", "the request this record executes; in ClinicalReport basedOn is "
                "explicitly the originating request", False),
    "originrequestid": ("OriginatingRequest", "certain", "the original prescription the record derives from - a mapping "
                        "no name match would find", False),
    "originatingrequest": ("OriginatingRequest", "certain", "named for the concept", False),

    # --- body site ----------------------------------------------------------
    "bodysite": ("BodyLocation", "likely", "FHIR's name for the body part concerned, which is what BodyLocation "
                 "names", True),
    "bodylocation": ("BodyLocation", "certain", "named for the concept", True),
    "bodylaterality": ("BodyLaterality", "certain", "named for the concept", True),
    "bodytopography": ("BodyTopography", "certain", "named for the concept", True),

    # --- product and device -------------------------------------------------
    "lotnumber": ("LotNumber", "certain", "manufacturer's batch identifier", True),
    "serialnumber": ("SerialNumber", "certain", "manufacturer's identifier of one device", True),
    "useddevice": ("UsedDevice", "certain", "named for the concept", True),
    "device": ("UsedDevice", "check", "the instrument the act was carried out with, per the subject-vs-instrument "
               "ruling; check any model where device is the subject of the act", True),
    "implantabledevice": ("ImplantableDevice", "certain", "named for the concept", True),
    "route": ("RouteOfAdministration", "certain", "the route a product is given by", True),
    "routeofadministration": ("RouteOfAdministration", "certain", "named for the concept", True),
    "administrationdate": ("AdministrationDate", "certain", "when the product was given", True),
    "vaccinationdate": ("AdministrationDate", "certain", "the date the vaccine was administered - AdministrationDate under "
                        "a vaccination-specific name", True),
    "occurrencedatetime": ("AdministrationDate", "check", "when the act happened; only AdministrationDate where the act "
                           "is giving a product", True),
}
CONFIDENCE_ORDER = {"certain": 0, "likely": 1, "check": 2}

DECISION_RE = re.compile(
    r"^```decision[ \t]+(?P<id>[A-Za-z0-9_.-]+)[ \t]*\n(?P<body>.*?)^```[ \t]*$",
    re.MULTILINE | re.DOTALL)


def norm(s):
    return re.sub(r"[^a-z0-9]", "", str(s or "").lower())


def sheet_rows(path):
    """(model, sheet, row, element name, current code, EN description)."""
    wb = openpyxl.load_workbook(path, data_only=True)
    model = os.path.splitext(os.path.basename(path))[0]
    if "Model" in wb.sheetnames:
        m = wb["Model"]
        for r in range(1, m.max_row + 1):
            if str(m.cell(r, 1).value or "").strip().lower() == "model name" and m.cell(r, 2).value:
                model = str(m.cell(r, 2).value).strip()
    out = []
    for name in wb.sheetnames:
        ws = wb[name]
        if ws.max_row < 3:
            continue
        h = {str(ws.cell(2, c).value or "").strip(): c for c in range(1, ws.max_column + 1)}
        if "Name" not in h or "Code" not in h:
            continue
        for r in range(3, ws.max_row + 1):
            element = ws.cell(r, h["Name"]).value
            if not element:
                continue
            desc = None
            for col in ("Description EN", "Short Label EN", "Description FR", "Short Label FR"):
                if h.get(col) and ws.cell(r, h[col]).value:
                    desc = str(ws.cell(r, h[col]).value)
                    break
            out.append({"model": model, "sheet": name, "row": r,
                        "element": str(element).strip(),
                        "code": (str(ws.cell(r, h["Code"]).value).strip()
                                 if ws.cell(r, h["Code"]).value else None),
                        "description": desc})
    return out


def collect():
    rows = []
    for p in sorted(glob.glob(os.path.join(ROOT, BOOKS, "*.xlsx"))):
        rows.extend(sheet_rows(p))
    return rows


def propose(rows):
    """Group unmapped elements by name and attach the rule that covers them."""
    groups = {}
    for row in rows:
        if row["code"]:
            continue
        leaf = row["element"].split(".")[-1]
        rule = RULES.get(norm(leaf)) or RULES.get(norm(row["element"]))
        if not rule:
            continue
        key = norm(leaf)
        code, confidence, why, nested_ok = rule
        if "." in row["element"] and not nested_ok:
            continue
        g = groups.setdefault(key, {"element": leaf, "code": code,
                                     "confidence": confidence, "why": why, "rows": []})
        g["rows"].append(row)
    return sorted(groups.values(),
                  key=lambda g: (CONFIDENCE_ORDER[g["confidence"]], -len(g["rows"]), g["element"]))



def suffix_of(element):
    """The element key the CSV carries.

    add_glossary_mappings.py matches an element by `suffix == key` or
    `suffix.endswith("." + key)`, so either the leaf or the full path works.
    The full path is what the existing rows use - reactions.note, not note -
    and it is the unambiguous one: a model can hold two nested `note`s.
    """
    return element


def read_mappings(path):
    """Existing rows, in file order. A row written before the Status column
    existed has no Status, and those 38 are all confirmed."""
    if not os.path.exists(path):
        return []
    rows = []
    with io.open(path, encoding="utf-8-sig", newline="") as fh:
        for r in csv.DictReader(fh, delimiter=";"):
            if not (r.get("Model") or "").strip():
                continue
            row = {k: (r.get(k) or "").strip() for k in FIELDS}
            # Write the implied Status out explicitly. Blank means confirmed to
            # every reader, but in a spreadsheet next to a column of 'proposed'
            # a blank cell reads as undecided, which is the opposite.
            row["Status"] = row["Status"] or "confirmed"
            rows.append(row)
    return rows


def write_mappings(path, rows):
    if os.path.exists(path):
        shutil.copy2(path, path + ".bak")
    with io.open(path, "w", encoding="utf-8-sig", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=FIELDS, delimiter=";",
                           extrasaction="ignore")
        w.writeheader()
        for r in rows:
            w.writerow(r)


def merge_into_csv(groups, existing):
    """Add proposals for anything not already in the CSV.

    A row already there is left exactly as it is, whatever its Status: a
    rejected mapping must not come back as a proposal on the next run, and a
    confirmed one must not be second-guessed.
    """
    seen = {(r["Model"], r["ElementSuffix"]) for r in existing}
    added = []
    terms = glossary_terms.load()
    drafted = glossary_terms.workbook_terms()

    # Re-check every row, not only the new ones. A term can be withdrawn or
    # renamed between runs, and a mapping to a term the glossary no longer
    # publishes should start showing that immediately.
    for r in existing:
        r["GlossaryStatus"] = glossary_terms.status_of(
            r["GlossaryCode"], terms, drafted)

    for g in groups:
        for row in g["rows"]:
            key = (row["model"], suffix_of(row["element"]))
            if key in seen:
                continue
            seen.add(key)
            added.append({
                "Model": row["model"],
                "ElementSuffix": suffix_of(row["element"]),
                "GlossaryCode": g["code"],
                "GlossaryStatus": glossary_terms.status_of(g["code"], terms, drafted),
                "Status": "proposed",
                "Confidence": g["confidence"],
                "Rationale": g["why"],
                "ElementDescription": (row["description"] or "")[:200],
            })
    added.sort(key=lambda r: (r["Model"], r["ElementSuffix"]))
    return existing + added



def suffix_of(element):
    """The element key the CSV carries.

    add_glossary_mappings.py matches an element by `suffix == key` or
    `suffix.endswith("." + key)`, so either the leaf or the full path works.
    The full path is what the existing rows use - reactions.note, not note -
    and it is the unambiguous one: a model can hold two nested `note`s.
    """
    return element


def read_mappings(path):
    """Existing rows, in file order. A row written before the Status column
    existed has no Status, and those 38 are all confirmed."""
    if not os.path.exists(path):
        return []
    rows = []
    with io.open(path, encoding="utf-8-sig", newline="") as fh:
        for r in csv.DictReader(fh, delimiter=";"):
            if not (r.get("Model") or "").strip():
                continue
            row = {k: (r.get(k) or "").strip() for k in FIELDS}
            # Write the implied Status out explicitly. Blank means confirmed to
            # every reader, but in a spreadsheet next to a column of 'proposed'
            # a blank cell reads as undecided, which is the opposite.
            row["Status"] = row["Status"] or "confirmed"
            rows.append(row)
    return rows


def write_mappings(path, rows):
    if os.path.exists(path):
        shutil.copy2(path, path + ".bak")
    with io.open(path, "w", encoding="utf-8-sig", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=FIELDS, delimiter=";", extrasaction="ignore")
        w.writeheader()
        for r in rows:
            w.writerow(r)


def merge_into_csv(groups, existing):
    """Add proposals for anything the CSV does not already carry.

    A row already there is left exactly as it is, whatever its Status: a
    rejected mapping must not come back as a proposal on the next run, and a
    confirmed one must not be second-guessed.
    """
    seen = {(r["Model"], r["ElementSuffix"]) for r in existing}
    added = []
    terms = glossary_terms.load()
    drafted = glossary_terms.workbook_terms()

    # Re-check every row, not only the new ones. A term can be withdrawn or
    # renamed between runs, and a mapping to a term the glossary no longer
    # publishes should start showing that immediately.
    for r in existing:
        r["GlossaryStatus"] = glossary_terms.status_of(
            r["GlossaryCode"], terms, drafted)

    for g in groups:
        for row in g["rows"]:
            key = (row["model"], suffix_of(row["element"]))
            if key in seen:
                continue
            seen.add(key)
            added.append({
                "Model": row["model"],
                "ElementSuffix": suffix_of(row["element"]),
                "GlossaryCode": g["code"],
                "GlossaryStatus": glossary_terms.status_of(g["code"], terms, drafted),
                "Status": "proposed",
                "Confidence": g["confidence"],
                "Rationale": g["why"],
                "ElementDescription": (row["description"] or "")[:200],
            })
    added.sort(key=lambda r: (r["Model"], r["ElementSuffix"]))
    return existing + added


def write_report(groups, rows, path):
    """A grouped read-through of the same proposals, for reviewing by concept
    rather than by model. The CSV is the artifact; this is a reading aid."""
    o = ["# Proposed glossary mappings\n",
         "- Generated: %s" % time.strftime("%Y-%m-%d %H:%M:%S"),
         "- Confirm or reject in `input/glossary_mappings.csv`; this file is "
         "only a grouped view of the same proposals.\n"]
    for level in ("certain", "likely", "check"):
        chunk = [g for g in groups if g["confidence"] == level]
        if not chunk:
            continue
        o.append("\n## %s (%d)\n" % (level.capitalize(), len(chunk)))
        for g in chunk:
            o.append("### `%s` -> **%s**   (%d model%s)\n"
                     % (g["element"], g["code"], len(g["rows"]),
                        "" if len(g["rows"]) == 1 else "s"))
            o.append("%s\n" % g["why"])
            o.append("| Model | Element | The model's own description |")
            o.append("|---|---|---|")
            for r in sorted(g["rows"], key=lambda r: r["model"]):
                d = (r["description"] or "").replace("|", "-").replace("\n", " ")
                o.append("| %s | `%s` | %s |" % (r["model"], r["element"], d[:110]))
            o.append("")

    unmatched = {}
    for r in rows:
        if r["code"]:
            continue
        leaf = suffix_of(r["element"])
        if RULES.get(norm(leaf)) or RULES.get(norm(r["element"])):
            continue
        unmatched.setdefault(leaf, []).append(r["model"])
    if unmatched:
        o.append("\n## No proposal (%d element names)\n" % len(unmatched))
        o.append("Nothing in the glossary obviously covers these. Most are "
                 "model-specific and correctly have no mapping; a few may be "
                 "worth a new glossary term.\n")
        o.append("| Element | Models |")
        o.append("|---|---|")
        for leaf, models in sorted(unmatched.items(), key=lambda kv: -len(kv[1]))[:60]:
            o.append("| `%s` | %d: %s |"
                     % (leaf, len(models), ", ".join(sorted(set(models))[:4])))
        o.append("")
    os.makedirs(os.path.dirname(path), exist_ok=True)
    io.open(path, "w", encoding="utf-8", newline="\n").write("\n".join(o))


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--mappings", default=MAPPINGS)
    ap.add_argument("--report", action="store_true",
                    help="also write a grouped read-through under glossary-changes/")
    ap.add_argument("--dry-run", action="store_true",
                    help="report what would be proposed, without writing the CSV")
    args = ap.parse_args()

    sys.stdout.reconfigure(encoding="utf-8")
    path = args.mappings if os.path.isabs(args.mappings) else os.path.join(ROOT, args.mappings)

    rows = collect()
    groups = propose(rows)
    existing = read_mappings(path)
    merged = merge_into_csv(groups, existing)
    added = len(merged) - len(existing)

    counts = {}
    for r in merged:
        key = r["Status"] or "confirmed"
        counts[key] = counts.get(key, 0) + 1

    print("Elements   : %d in %d workbook(s), %d already mapped"
          % (len(rows), len(set(r["model"] for r in rows)),
             sum(1 for r in rows if r["code"])))
    print("Proposals  : %d new row(s) in %d decision group(s)" % (added, len(groups)))
    for level in ("certain", "likely", "check"):
        n = sum(1 for g in groups if g["confidence"] == level)
        if n:
            print("             %-8s %2d group(s)" % (level, n))
    if args.dry_run:
        print("\n--dry-run: %s not written" % args.mappings)
        return 0

    write_mappings(path, merged)
    unapproved = {}
    for r in merged:
        if r["GlossaryStatus"] and r["GlossaryStatus"] != "approved":
            unapproved.setdefault((r["GlossaryCode"], r["GlossaryStatus"]), 0)
            unapproved[(r["GlossaryCode"], r["GlossaryStatus"])] += 1
    if unapproved:
        print("\nTargets that are not approved glossary terms:")
        for (code, status), n in sorted(unapproved.items()):
            print("             %-24s %-16s %3d row(s)" % (code, status, n))
        print("             merge_mappings_to_xlsx.py will not write these "
              "without --allow-unapproved")

    print("\nCSV        : %s" % args.mappings)
    print("             %s" % "  ".join("%s=%d" % kv for kv in sorted(counts.items())))
    if args.report:
        rp = os.path.join(ROOT, OUT_DIR,
                          "model-mappings-%s.md" % time.strftime("%Y%m%d_%H%M%S"))
        write_report(groups, rows, rp)
        print("Read-through: %s" % os.path.relpath(rp, ROOT))
    print("\nNext: set Status in the CSV (proposed -> confirmed or rejected),")
    print("      then: python merge_mappings_to_xlsx.py")
    return 0


if __name__ == "__main__":
    sys.exit(main())
