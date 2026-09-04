"""
The glossary terms a mapping is allowed to point at.

A mapping is only meaningful if its target is a term the glossary actually
publishes and has approved. The clinical glossary holds 24 approved terms and
59 more still without a status, so pointing at an unapproved one is an easy
mistake to make by hand - and one nothing would otherwise catch until the term
failed to resolve on the site.

Used by propose_model_mappings.py and merge_mappings_to_xlsx.py so both judge a
target the same way.
"""

import csv
import glob
import io
import json
import os
import warnings

ROOT = os.path.dirname(os.path.abspath(__file__))
SOURCES = [os.path.join("input", "ClinicalGlossary.csv"),
           os.path.join("input", "OperationalGlossary.csv")]

# The published CodeSystems, whose canonicals a mapping's Coding.system must
# carry. Read from the generated resources rather than hardcoded here, so a
# change to a canonical reaches the models without a second edit.
CODESYSTEMS = [os.path.join("_resources", "glossary", "CodeSystem-glossary.json"),
               os.path.join("_resources", "glossary",
                            "CodeSystem-operational-glossary.json")]

# The statuses that make a term usable as a mapping target. `accepted` is what
# import_glossary_xlsx.py writes for a workbook row marked active; `active` is
# what the hand-maintained operational glossary uses for the same thing.
APPROVED = {"accepted", "active"}

NOT_IN_GLOSSARY = "not in glossary"
DRAFTED = "in the workbook, not approved"


def load(sources=None):
    """code -> status, across every glossary CSV.

    These are the generated CSVs rather than the workbook: they are what the
    site publishes, so a term absent from them is absent from the glossary
    whatever the workbook says.
    """
    terms = {}
    for rel in (sources or SOURCES):
        path = rel if os.path.isabs(rel) else os.path.join(ROOT, rel)
        if not os.path.exists(path):
            continue
        with io.open(path, encoding="utf-8-sig", newline="") as fh:
            for r in csv.DictReader(fh, delimiter=";"):
                code = (r.get("Term") or "").strip()
                if code:
                    terms[code] = (r.get("Status") or "").strip().lower()
    return terms


def workbook_terms():
    """Every term in the glossary workbook, approved or not.

    The generated CSVs carry only the approved terms, so a mapping to one still
    being drafted reads there as if the term did not exist. Those are different
    problems with different fixes - approve the term, versus the term is a
    typo - and telling them apart is worth reading the workbook for.

    Returns an empty set rather than failing: this is diagnosis, and the
    glossary CSVs remain the authority on what is approved.
    """
    hits = sorted(glob.glob(os.path.join(ROOT, "input", "Glossaire CareSets*.xlsx")))
    if not hits:
        return set()
    try:
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            import openpyxl
            wb = openpyxl.load_workbook(hits[0], data_only=True, read_only=True)
    except Exception:
        return set()
    codes = set()
    for ws in wb.worksheets:
        head = {str(ws.cell(1, c).value or "").strip(): c
                for c in range(1, (ws.max_column or 0) + 1)}
        if "Item" not in head:
            continue
        for r in range(2, (ws.max_row or 0) + 1):
            label = ws.cell(r, head["Item"]).value
            if label:
                # The same derivation import_glossary_xlsx.py uses: the label is
                # spaced Title Case, the code is that with the spaces removed.
                codes.add("".join(w[:1].upper() + w[1:]
                                  for w in str(label).split()))
    wb.close()
    return codes


def systems():
    """code -> the CodeSystem canonical that publishes it.

    A term belongs to one of two CodeSystems, the clinical glossary or the
    operational one, and a Coding naming the wrong system does not resolve.
    Reading the generated CodeSystems keeps the answer right without this
    module having to know which glossary holds which term.
    """
    out = {}
    for rel in CODESYSTEMS:
        path = rel if os.path.isabs(rel) else os.path.join(ROOT, rel)
        if not os.path.exists(path):
            continue
        try:
            doc = json.load(io.open(path, encoding="utf-8"))
        except ValueError:
            continue
        url = doc.get("url")
        if not url:
            continue
        for concept in doc.get("concept", []):
            if concept.get("code"):
                out[concept["code"]] = url
    return out


def status_of(code, terms, drafted=None):
    """The mapping target's standing, as it should be recorded in the CSV.

    One of:

      approved                      the glossary publishes it and it is approved
      in the workbook, not approved  the term is being drafted; approve it, or
                                     point the mapping elsewhere
      not in glossary                no such term - usually a typo
      <the term's own status>        published but not in an approved state

    Anything other than `approved` is a flag for a reviewer, not an error: a
    mapping to a term still being drafted is a legitimate thing to have
    proposed, as long as it is visible.

    Pass `drafted` (from workbook_terms()) when checking many codes, so the
    workbook is opened once rather than per code.
    """
    code = (code or "").strip()
    if not code:
        return ""
    if code not in terms:
        if drafted is None:
            drafted = workbook_terms()
        return DRAFTED if code in drafted else NOT_IN_GLOSSARY
    status = terms[code]
    return "approved" if status in APPROVED else (status or "no status")


def is_approved(code, terms, drafted=None):
    return status_of(code, terms, drafted) == "approved"
