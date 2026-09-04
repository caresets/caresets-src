"""
Generate FHIR StructureDefinitions from the logical-model workbooks.

Model authors write the workbook in French and Dutch; English is added and
verified before publication. This turns a workbook into the StructureDefinition
that goes off to be published:

  models/xls/<Model>.xlsx          authored FR + NL, English added here
        |  import_logical_model_xlsx.py     <- this script
        v
  models/generated/StructureDefinition-<Model>.json
        |  (hand off to the publication process)
        v
  ...published as part of the eHealth package...
        |  import_models_zip.py             <- how they come back
        v
  input/models/   ->  build_content.py  ->  _resources/models/

Note the direction. `input/models/` holds models that have BEEN published and
came back in an export; it is not written by this script. Generating straight
into it would put unpublished output where published models live, and the next
export would overwrite it with no warning.

English is the base language: `short` and `definition` carry the English text,
and French and Dutch travel in the standard FHIR `translation` extension on the
matching `_short` / `_definition`. Where a language is missing the base falls
back to whichever text exists - French, then Dutch - so a model that has not
been translated yet still publishes, in the language it was written in.

The workbook owns the element names, the three languages, cardinality, data
type and ValueSet binding. Everything else in an existing model - publisher,
contact, jurisdiction, narrative, dates - is preserved, because a spreadsheet
is not the right place to keep it and regenerating it would lose it.

Usage:
  python import_logical_model_xlsx.py                          # every workbook in models/xls
  python import_logical_model_xlsx.py --model BeModelVaccination
  python import_logical_model_xlsx.py --dry-run
"""

import argparse
import copy
import csv
import glob
import io
import json
import os
import sys
import unicodedata
import warnings

warnings.filterwarnings("ignore")
import openpyxl  # noqa: E402

import glossary_terms  # noqa: E402

ROOT = os.path.dirname(os.path.abspath(__file__))
TRANSLATION_URL = "http://hl7.org/fhir/StructureDefinition/translation"
LANGS = ("en", "fr", "nl")          # order matters: the base falls back this way
FIRST_DATA_ROW = 3
DEFAULT_IN = os.path.join("models", "xls")
DEFAULT_OUT = os.path.join("models", "generated")

COLUMNS = {
    "name": "name",
    "short label fr": "short_fr", "description fr": "description_fr",
    "short label nl": "short_nl", "description nl": "description_nl",
    "short label en": "short_en", "description en": "description_en",
    "data type": "datatype",
    "min occurrence": "min", "max occurrence": "max",
    "valueset": "valueset",
    "binding strength": "binding_strength",
    "code": "glossary_code",
    "relationship": "glossary_relationship",
}


def norm(s):
    s = unicodedata.normalize("NFKD", str(s or ""))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return " ".join(s.lower().split())


def clean(v):
    if v is None:
        return None
    s = str(v).replace("\xa0", " ").strip()
    return s or None


def set_text(owner, field, texts):
    """Write one field as an English base plus translations.

    Anything already on `_field` that is not a translation - other extensions -
    is left alone, so this can be run over a model repeatedly.
    """
    present = {k: v for k, v in texts.items() if v}
    if not present:
        owner.pop(field, None)
        owner.pop("_" + field, None)
        return None

    base_lang = next((l for l in LANGS if present.get(l)), None)
    owner[field] = present[base_lang]

    others = [(l, t) for l, t in present.items() if l != base_lang]
    sibling = owner.get("_" + field) or {}
    keep = [e for e in sibling.get("extension", []) or [] if e.get("url") != TRANSLATION_URL]
    exts = keep + [
        {
            "extension": [
                {"url": "lang", "valueCode": lang},
                {"url": "content", "valueString": text},
            ],
            "url": TRANSLATION_URL,
        }
        for lang, text in sorted(others)
    ]
    if exts:
        owner["_" + field] = {"extension": exts}
    else:
        owner.pop("_" + field, None)
    return base_lang


def read_model_sheet(wb):
    """The Model sheet, in either shape it can take.

    The template hands editors a TABLE - one row per data set, so several
    models can live in one workbook. A workbook produced by
    export_logical_model_xlsx.py holds a single model and writes the same
    fields as KEY/VALUE pairs down two columns. Both are accepted, and both
    return {sheet name or None: {field: value}}.
    """
    if "Model" not in wb.sheetnames:
        return {}
    ws = wb["Model"]
    first = norm(ws.cell(1, 1).value)

    if first == "data set sheet":
        headings = {}
        for c in range(1, ws.max_column + 1):
            h = norm(ws.cell(1, c).value)
            if h:
                headings[c] = h
        out = {}
        for r in range(2, ws.max_row + 1):
            row = {h: clean(ws.cell(r, c).value) for c, h in headings.items()}
            sheet = row.pop("data set sheet", None)
            if any(row.values()):
                out[sheet] = row
        return out

    single = {}
    for r in range(1, ws.max_row + 1):
        k, v = clean(ws.cell(r, 1).value), clean(ws.cell(r, 2).value)
        if k:
            single[norm(k)] = v
    return {None: single} if single else {}


def element_sheets(wb):
    """Every sheet that holds data elements, in workbook order.

    Identified by having a `Name` column in row 2, rather than by excluding
    known sheet names. ValueSet tabs are not reliably prefixed "VS" - the
    exporter names them after the binding, e.g. be-vs-allergyintolerance-type -
    so an exclusion list both misses them and would mistake one for a model if
    it happened to carry the right header.
    """
    out = []
    for n in wb.sheetnames:
        ws = wb[n]
        if ws.max_row < 2:
            continue
        if any(norm(ws.cell(2, c).value) == "name"
               for c in range(1, min(ws.max_column, 30) + 1)):
            out.append(n)
    return out


def read_elements(wb, sheet_name):
    """The element rows of one data-set sheet."""
    ws = wb[sheet_name]
    idx = {}
    for c in range(1, ws.max_column + 1):
        key = COLUMNS.get(norm(ws.cell(2, c).value))
        if key and key not in idx:
            idx[key] = c
    if "name" not in idx:
        return [], ws.title

    rows = []
    for r in range(FIRST_DATA_ROW, ws.max_row + 1):
        def cell(key):
            c = idx.get(key)
            return clean(ws.cell(r, c).value) if c else None
        name = cell("name")
        if not name:
            continue
        rows.append({k: cell(k) for k in idx})
    return rows, ws.title


def build(existing, model, rows, model_name, systems=None):
    """The StructureDefinition, preserving whatever the workbook does not own."""
    doc = copy.deepcopy(existing) if existing else {
        "resourceType": "StructureDefinition",
        "fhirVersion": "4.0.1",
        "kind": "logical",
        "abstract": False,
        "derivation": "specialization",
        "baseDefinition": "http://hl7.org/fhir/StructureDefinition/Base",
        "status": "draft",
    }
    name = model.get("model name") or model_name
    url = model.get("canonical url") or doc.get("url")

    # The element paths are rooted on the model's own root element path, which
    # is not always the `name`: BeModelVaccination's elements sit under
    # be-model-vaccination, and MedicationLine's under BeModelMedicationLine.
    # Rebuilding paths from `name` would rename every element in those models.
    root_path = name
    for e in (existing or {}).get("differential", {}).get("element", []):
        if "." not in e.get("path", "."):
            root_path = e["path"]
            break
    else:
        if existing and existing.get("type"):
            root_path = existing["type"].rstrip("/").split("/")[-1]

    # `id` identifies the resource; the workbook does not own it, so an
    # existing one is kept rather than renamed to match the display name.
    doc["id"] = (existing or {}).get("id") or name
    doc["name"] = name
    if url:
        doc["url"] = url
        doc["type"] = url
    if model.get("version"):
        doc["version"] = model["version"]
    if model.get("status"):
        doc["status"] = model["status"]

    set_text(doc, "title", {l: model.get("title " + l) for l in LANGS})
    base_lang = set_text(doc, "description", {l: model.get("description " + l) for l in LANGS})
    # `language` states what the untranslated fields are written in, so a
    # reader never has to guess which language the base text is.
    doc["language"] = base_lang or "en"

    # The root element is not a row in the sheet, and its short/definition are
    # not always the model's title/description - so an existing root is carried
    # over untouched rather than overwritten from the Model sheet.
    root = None
    for e in (existing or {}).get("differential", {}).get("element", []):
        if "." not in e.get("path", "."):
            root = copy.deepcopy(e)
            break
    if root is None:
        root = {"id": root_path, "path": root_path}
        set_text(root, "short", {l: model.get("title " + l) for l in LANGS})
        set_text(root, "definition", {l: model.get("description " + l) for l in LANGS})

    systems = systems or {}

    elements = [root]
    for row in rows:
        path = "%s.%s" % (root_path, row["name"])
        e = {"id": path, "path": path}
        set_text(e, "short", {l: row.get("short_" + l) for l in LANGS})
        set_text(e, "definition", {l: row.get("description_" + l) for l in LANGS})
        if row.get("min") is not None:
            try:
                e["min"] = int(row["min"])
            except (TypeError, ValueError):
                pass
        if row.get("max"):
            e["max"] = str(row["max"])
        if row.get("datatype"):
            e["type"] = [{"code": t.strip()} for t in row["datatype"].split("|") if t.strip()]
        if row.get("valueset"):
            # FHIR's binding strengths, weakest last. A blank column keeps the
            # previous behaviour rather than silently tightening a binding.
            strength = norm(row.get("binding_strength")) or "preferred"
            if strength not in ("required", "extensible", "preferred", "example"):
                print("  ! %s.%s: unknown binding strength %r, using preferred"
                      % (root_path, row["name"], row.get("binding_strength")))
                strength = "preferred"
            e["binding"] = {"strength": strength, "valueSet": row["valueset"]}
        if row.get("glossary_code"):
            # The workbook Code column is the mapping, so the generated
            # model carries it too. Without this the codes reached only the
            # site's own copies, and the StructureDefinitions handed over
            # for publication went out with none.
            code = row["glossary_code"].strip()
            system = systems.get(code)
            if system:
                e["code"] = [{"system": system, "code": code}]
            else:
                print("  ! %s.%s: %s is in no published CodeSystem, "
                      "code omitted" % (root_path, row["name"], code))
        elements.append(e)

    doc["differential"] = {"element": elements}
    # A snapshot is the publisher's job; a stale one is worse than none, so a
    # regenerated model carries only the differential.
    doc.pop("snapshot", None)
    doc.pop("text", None)
    return doc


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--in-dir", dest="in_dir", default=DEFAULT_IN)
    ap.add_argument("--out-dir", dest="out_dir", default=DEFAULT_OUT,
                    help="where to write the generated StructureDefinitions "
                         "(default: models/generated - NOT input/models, which "
                         "holds what came back from publication)")
    ap.add_argument("--baseline", default=os.path.join("input", "models"),
                    help="already-published models to merge into, so publisher, "
                         "contact, jurisdiction and canonical URL are preserved "
                         "(default: input/models)")
    ap.add_argument("--model", nargs="+", metavar="NAME")
    ap.add_argument("--mappings-out", dest="mappings_out",
                    default=os.path.join("input", "glossary_mappings.csv"),
                    help="where to write the model-to-glossary mappings read from the "
                         "workbooks (default: input/glossary_mappings.csv). Skipped when "
                         "--model limits the run, since that would write a partial file.")
    ap.add_argument("--no-mappings", action="store_true",
                    help="do not write the mappings file")
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    os.makedirs(os.path.join(ROOT, args.out_dir), exist_ok=True)
    books = sorted(glob.glob(os.path.join(ROOT, args.in_dir, "*.xlsx")))
    if not books:
        sys.exit("No workbooks in %s" % args.in_dir)

    # Which CodeSystem publishes each glossary term, for the Coding on
    # element.code. A term in neither is reported per element rather than
    # written with a guessed system, which would not resolve.
    code_systems = glossary_terms.systems()

    # Merge into the last published version of each model, if there is one, so
    # metadata the workbook does not carry survives a regeneration.
    existing_by_name = {}
    for p in glob.glob(os.path.join(ROOT, args.baseline, "**", "*.json"), recursive=True):
        try:
            d = json.load(io.open(p, encoding="utf-8"))
        except ValueError:
            continue
        if d.get("resourceType") == "StructureDefinition" and d.get("name"):
            existing_by_name[d["name"]] = p

    sys.stdout.reconfigure(encoding="utf-8")
    written = skipped = 0
    mappings = []
    for book in books:
        stem = os.path.splitext(os.path.basename(book))[0]
        wb = openpyxl.load_workbook(book, data_only=True)
        models = read_model_sheet(wb)
        sheets = element_sheets(wb)
        if not sheets:
            print("  ! %s: no data-set sheet" % os.path.basename(book))
            continue

        # A workbook may hold several data sets, one sheet each - which is the
        # shape of the template handed to editors. Each sheet is matched to its
        # row on the Model sheet; a single-model workbook keys on None.
        for sheet in sheets:
            model = models.get(sheet)
            if model is None:
                model = models.get(None, {}) if len(sheets) == 1 else {}
            rows, _ = read_elements(wb, sheet)
            for row in rows:
                if row.get("glossary_code"):
                    mappings.append((model.get("model name") or sheet, row["name"],
                                     row["glossary_code"]))
            name = model.get("model name") or (stem if len(sheets) == 1 else sheet)
            if not rows:
                print("  - %-32s no data elements, skipped" % name)
                continue
            if args.model and name not in args.model:
                continue

            prev_path = existing_by_name.get(name)
            prev = json.load(io.open(prev_path, encoding="utf-8")) if prev_path else None
            doc = build(prev, model, rows, name, code_systems)

            out = os.path.join(ROOT, args.out_dir,
                               os.path.basename(prev_path) if prev_path
                               else "StructureDefinition-%s.json" % name)
            langs = set()
            for e in doc["differential"]["element"]:
                for f in ("short", "definition"):
                    for ext in (e.get("_" + f) or {}).get("extension", []) or []:
                        for sub in ext.get("extension", []) or []:
                            if sub.get("url") == "lang":
                                langs.add(sub.get("valueCode"))
            print("  %-34s %2d element(s), base %s, translations: %s%s"
                  % (name, len(rows), doc.get("language"),
                     ", ".join(sorted(langs)) or "none",
                     "" if prev_path else "   [NEW]"))
            if args.dry_run:
                skipped += 1
                continue
            with io.open(out, "w", encoding="utf-8", newline="\n") as fh:
                json.dump(doc, fh, ensure_ascii=False, indent=2)
            written += 1

    print()
    if args.dry_run:
        print("--dry-run: %d model(s) inspected, nothing written" % skipped)
        if mappings:
            print("would write %d mapping(s) to %s" % (len(mappings), args.mappings_out))
        return 0

    print("Wrote %d model(s) into %s" % (written, args.out_dir))

    # The Code column is where analysts author the model-to-glossary mapping,
    # so the CSV is generated from the workbooks rather than kept by hand. A
    # run limited to some models would only see part of the picture, so it is
    # skipped rather than writing a file that drops everything else.
    if args.no_mappings:
        pass
    elif args.model:
        print("mappings not written: --model limits the run to part of the set")
    else:
        path = args.mappings_out if os.path.isabs(args.mappings_out) \
            else os.path.join(ROOT, args.mappings_out)
        fields = ["Model", "ElementSuffix", "GlossaryCode", "Status",
                  "Confidence", "Rationale", "ElementDescription"]

        # The workbooks are the source for a confirmed mapping, but the CSV
        # also carries proposals awaiting review and mappings already
        # rejected. Regenerating from the workbooks alone would silently
        # discard both - and a rejected row coming back as a fresh proposal
        # on the next run would put a settled question to the reviewer again.
        pending = {}
        if os.path.exists(path):
            with io.open(path, encoding="utf-8-sig", newline="") as fh:
                for r in csv.DictReader(fh, delimiter=";"):
                    key = ((r.get("Model") or "").strip(),
                           (r.get("ElementSuffix") or "").strip())
                    status = (r.get("Status") or "confirmed").strip().lower()
                    if all(key) and status != "confirmed":
                        pending[key] = {f: (r.get(f) or "").strip()
                                        for f in fields}

        seen, rows = set(), []
        for model, suffix, code in mappings:
            key = (model, suffix)
            if key in seen:
                continue
            seen.add(key)
            # A mapping now in a workbook was confirmed by being written
            # there, so it supersedes any pending row for the same element.
            pending.pop(key, None)
            rows.append({"Model": model, "ElementSuffix": suffix,
                         "GlossaryCode": code, "Status": "confirmed"})
        rows.extend(pending.values())
        rows.sort(key=lambda r: (r["Model"], r["ElementSuffix"]))

        with io.open(path, "w", encoding="utf-8-sig", newline="") as fh:
            w = csv.DictWriter(fh, fieldnames=fields, delimiter=";",
                               lineterminator="\n", extrasaction="ignore")
            w.writeheader()
            for r in rows:
                w.writerow(r)
        kept = len(pending)
        print("Wrote %d mapping(s) to %s%s"
              % (len(rows), args.mappings_out,
                 " (%d awaiting review, carried over)" % kept if kept else ""))
    return 0


if __name__ == "__main__":
    sys.exit(main())
