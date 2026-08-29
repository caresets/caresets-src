"""
Convert the glossary workbook into the CSV that generate_glossary.py reads.

The workbook is the source. This turns it into `input/ClinicalGlossary.csv`,
which build_content.py then renders into the published CodeSystems, so the
usual flow is: edit the workbook -> run build_content.py -> the site is updated.

  input/Glossaire CareSets*.xlsx  (sheet "Glossaire v1")
        |  import_glossary_xlsx.py      <- this script
        v
  input/ClinicalGlossary.csv
        |  generate_glossary.py
        v
  _resources/glossary/CodeSystem-glossary.json

Definition and Description stay separate all the way through: `Definition XX`
becomes the term's text in language XX, and `Description XX` becomes the
`XX Note` column. Merging them, as the CSV used to, is what made a note to
entry read as part of the definition.

Two things this script deliberately does NOT touch:

  * `input/OperationalGlossary.csv` - the workbook's "Glossaire operationnel"
    sheet holds 2 rows against that file's 13 terms, so the operational
    glossary has no workbook source and is still maintained by hand.
  * Any term already in the CSV but absent from the workbook: it is kept and
    reported, never silently dropped. Use --prune to remove them deliberately.

Usage:
  python import_glossary_xlsx.py                 # convert, then diff-report
  python import_glossary_xlsx.py --dry-run       # report without writing
  python import_glossary_xlsx.py --prune         # also drop terms not in the workbook
"""

import argparse
import csv
import glob
import io
import os
import sys
import unicodedata
import warnings

warnings.filterwarnings("ignore")
import openpyxl  # noqa: E402

ROOT = os.path.dirname(os.path.abspath(__file__))
DEFAULT_SHEET = "Glossaire v1"
DEFAULT_CSV = os.path.join("input", "ClinicalGlossary.csv")
FIELDNAMES = ["Term", "Display", "Status", "Synonym", "FR", "EN", "NL",
              "FR Note", "EN Note", "NL Note"]

# Workbook header -> what we call it. Matching folds case and accents, so
# "Definition NL" and "Définition NL" are the same column.
COLUMNS = {
    "statut de la def": "status",
    "item": "term",
    "synonym": "synonym",
    "synonyme": "synonym",
    "definition fr": "fr",
    "description fr": "fr_note",
    "definition nl": "nl",
    "description nl": "nl_note",
    "definition en": "en",
    "description en": "en_note",
}

# The workbook records editorial state; the CodeSystem records lifecycle.
STATUS = {"active": "accepted", "draft": "draft", "proposed": "proposed",
          "deprecated": "deprecated", "rejected": "rejected"}


def code_of(label):
    """The CodeSystem code for a workbook item.

    The Item column is the human label, in spaced Title Case ("Business
    Identifier"); the code is that with the spaces removed
    ("BusinessIdentifier"). Keeping them distinct is what lets the sheet read
    naturally without changing the codes that glossary_mappings.csv and every
    published CodeSystem already point at.
    """
    return "".join(str(label or "").split())


def norm(s):
    s = unicodedata.normalize("NFKD", str(s or ""))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return " ".join(s.lower().split())


def clean(v):
    """Cell text as it should land in the CSV: no stray whitespace, no hard
    spaces, and newlines flattened so a row stays a row."""
    if v is None:
        return ""
    s = str(v).replace("\xa0", " ").replace("\r\n", "\n").strip()
    return " ".join(s.split())


def find_workbook(explicit):
    if explicit:
        return explicit
    hits = sorted(glob.glob(os.path.join(ROOT, "input", "Glossaire CareSets*.xlsx")))
    if not hits:
        sys.exit("No workbook found in input/. Pass --workbook.")
    if len(hits) > 1:
        sys.exit("More than one workbook in input/, so the source is ambiguous:\n  %s\n"
                 "Keep one, or pass --workbook." % "\n  ".join(os.path.basename(h) for h in hits))
    return hits[0]


def read_workbook(path, sheet, status_filter):
    wb = openpyxl.load_workbook(path, data_only=True)
    if sheet not in wb.sheetnames:
        sys.exit("Sheet %r not in %s (has: %s)" % (sheet, path, wb.sheetnames))
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        sys.exit("Sheet %r is empty" % sheet)

    idx = {}
    for i, h in enumerate(rows[0]):
        key = COLUMNS.get(norm(h))
        if key and key not in idx:
            idx[key] = i
    for required in ("term", "status", "fr"):
        if required not in idx:
            sys.exit("Sheet %r has no %r column (found: %s)"
                     % (sheet, required, sorted(idx)))

    out, seen = [], {}
    for n, r in enumerate(rows[1:], start=2):
        def cell(key):
            i = idx.get(key)
            return clean(r[i]) if i is not None and i < len(r) else ""

        term, status = cell("term"), cell("status")
        if not term or not status:
            continue
        if status_filter and norm(status) not in [norm(s) for s in status_filter]:
            continue
        if norm(term) in seen:
            print("  ! row %d: %r duplicates row %d; keeping the first"
                  % (n, term, seen[norm(term)]))
            continue
        seen[norm(term)] = n
        out.append({
            "Term": code_of(term),
            "Display": term,
            "Status": STATUS.get(norm(status), norm(status)),
            "Synonym": cell("synonym"),
            "FR": cell("fr"), "EN": cell("en"), "NL": cell("nl"),
            "FR Note": cell("fr_note"), "EN Note": cell("en_note"),
            "NL Note": cell("nl_note"),
        })
    return out, sorted(idx)


def read_csv(path):
    if not os.path.exists(path):
        return []
    with io.open(path, encoding="utf-8-sig", newline="") as fh:
        return [r for r in csv.DictReader(fh, delimiter=";") if (r.get("Term") or "").strip()]


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--workbook", help="default: the single input/Glossaire CareSets*.xlsx")
    ap.add_argument("--sheet", default=DEFAULT_SHEET)
    ap.add_argument("--out", default=DEFAULT_CSV)
    ap.add_argument("--status", nargs="*", default=["Active"],
                    help="workbook statuses to publish (default: Active); "
                         "pass no value to take every row")
    ap.add_argument("--prune", action="store_true",
                    help="drop CSV terms that are absent from the workbook "
                         "(default: keep them and report)")
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    wb_path = find_workbook(args.workbook)
    rows, found = read_workbook(wb_path, args.sheet, args.status)
    if not rows:
        sys.exit("No rows matched status %s in %s" % (args.status, args.sheet))

    out_path = args.out if os.path.isabs(args.out) else os.path.join(ROOT, args.out)
    existing = read_csv(out_path)
    old = {norm(r["Term"]): r for r in existing}
    new = {norm(r["Term"]): r for r in rows}

    # Terms the workbook does not cover stay unless --prune, so a sheet that is
    # only partly filled in cannot quietly delete published concepts.
    kept = []
    if not args.prune:
        for key, r in old.items():
            if key not in new:
                kept.append(r.get("Term"))
                merged = {f: (r.get(f) or "") for f in FIELDNAMES}
                merged["Term"] = r.get("Term")
                rows.append(merged)

    rows.sort(key=lambda r: norm(r["Term"]))

    added = [r["Term"] for k, r in new.items() if k not in old]
    changed = []
    for k, r in new.items():
        if k not in old:
            continue
        before = old[k]
        for f in ("FR", "EN", "NL", "Status", "Synonym"):
            if (before.get(f) or "").strip() != (r.get(f) or "").strip():
                changed.append((r["Term"], f))
                break

    sys.stdout.reconfigure(encoding="utf-8")
    print("Workbook: %s" % os.path.relpath(wb_path, ROOT).replace("\\", "/"))
    print("Sheet   : %s  (columns found: %s)" % (args.sheet, ", ".join(found)))
    print("Status  : %s" % (args.status or "every row"))
    print()
    print("%d term(s) from the workbook, %d already in the CSV" % (len(new), len(old)))
    print("  added   %d%s" % (len(added), ": " + ", ".join(added) if added else ""))
    print("  changed %d%s" % (len(changed),
                              ": " + ", ".join("%s (%s)" % c for c in changed[:8]) if changed else ""))
    if kept:
        print("  kept    %d not in the workbook (use --prune to drop): %s"
              % (len(kept), ", ".join(kept)))
    relabelled = [(r["Display"], r["Term"]) for r in new.values()
                  if r.get("Display") and r["Display"] != r["Term"]]
    if relabelled:
        print("  label != code for %d term(s), e.g. %s"
              % (len(relabelled), ", ".join("%s -> %s" % x for x in relabelled[:4])))
    missing_en = [r["Term"] for r in rows if not r.get("EN")]
    if missing_en:
        print("  ! no EN for %d term(s): %s" % (len(missing_en), ", ".join(missing_en)))

    if args.dry_run:
        print("\n--dry-run: %s not written" % args.out)
        return 0

    with io.open(out_path, "w", encoding="utf-8-sig", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=FIELDNAMES, delimiter=";", lineterminator="\n")
        w.writeheader()
        w.writerows(rows)
    print("\nWrote %s (%d rows)" % (args.out, len(rows)))
    return 0


if __name__ == "__main__":
    sys.exit(main())
